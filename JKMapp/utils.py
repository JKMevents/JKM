import qrcode

# genrate qr code function 
from PIL import Image, ImageDraw, ImageFont
import datetime
import os
from openpyxl import Workbook, load_workbook
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
#import os

import os.path

import cv2
from pyzbar.pyzbar import decode


COUNTERS_FILE_PATH = 'JKMapp/static/counter_info.xlsx'



def generate_multiple_qr_codes(num_qr):
    qr_codes = []
    for i in range(1, num_qr+1):
        qr = qrcode.QRCode(
            version=1,
            box_size=5,
            border=4,
        )
        qr_data = f'JKM2024_{i}'
        qr.add_data(qr_data)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")
        qr_codes.append(qr_img)
    return qr_codes


def combine_qr_codes(qr_codes, margin=10):
    qr_width = qr_codes[0].size[0]
    qr_height = qr_codes[0].size[1]
    total_width = qr_width
    total_height = len(qr_codes) * qr_height + (len(qr_codes) - 1) * margin
    combined_image = Image.new('RGB', (total_width, total_height), color='white')
    y_offset = 0
    for qr in qr_codes:
        combined_image.paste(qr, (0, y_offset))
        y_offset += qr_height + margin
    return combined_image


def create_canvas(number, ticket_template, margin =10):
    ticket_width, ticket_height = ticket_template.size

    # Calculate total dimensions for combined image
    total_width = ticket_width

    total_height = (int(number) * ticket_height) +  margin

    # Create a new image with dimensions for combined tickets
    ticket_canvas = Image.new('RGB', (total_width, total_height), color='white')
    return ticket_canvas

def generate_qr_codes(num_qr=1, data_prefix="JKM2024"):
    qr_codes = []
    for i in range(1, num_qr+1):
        qr = qrcode.QRCode(
            version=1,
            box_size=10,
            border=4,
        )
        qr_data = f'{data_prefix}_{i}'
        qr.add_data(qr_data)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")
        qr_codes.append(qr_img)
    return qr_codes


def put_tickets(canvas, number, qr_codes, ticket_template, margin=10):
        # Paste each QR code on a ticket layout
        y_offset = 0
        ticket_width, ticket_height = ticket_template.size

        for qr in qr_codes:
            # Paste ticket layout on combined image
            canvas.paste(ticket_template, (0, y_offset))
            y_offset += ticket_height + margin

        return canvas


def putqr(canvas, qr_codes, ticket_template ,margin = 10):
    y_offset = 0
    ticket_width, ticket_height = ticket_template.size
    qr_width, qr_height = qr_codes[0].size
    for qr in qr_codes:
        # Paste QR code on ticket layout
        qr_x_offset = ticket_width - qr_width - margin
        qr_y_offset = y_offset + (ticket_height - qr_height) // 2
        canvas.paste(qr, (qr_x_offset, qr_y_offset))
        y_offset += ticket_height + margin
    return canvas

def add_text_to_image(canvas, number, ticket_template, text_info, font_path, font_size, color, margin=10):

    ticket_width, ticket_height = ticket_template.size

    # Create a draw object.
    draw = ImageDraw.Draw(canvas)

    # Create a font object.
    font = ImageFont.truetype(font_path, size=font_size)

    # Draw the text on the new image.
    for a in range(number):
        for text, position in text_info:
            draw.text((position[0], position[1] + ((ticket_height)+margin)*a ), text, fill=color, font=font)
            
            # Adjust y_offset for next text line

  # Adjust y_offset for next ticket

    # Return the new image with the text added.
    return canvas



def create_tickets(number, start_number, ticket_template, margin=10, data_prefix="JKM2024", font_path="JKMapp/static/Roboto-Medium.ttf"):
    # Define prerequisites and variables
    font_size = 60
    color = (255, 0, 0)
    text_info = []

    # Get current date and time
    current_datetime = datetime.datetime.utcnow() + datetime.timedelta(hours=5, minutes=30)
    formatted_date = current_datetime.strftime("%d-%m-%Y")
    formatted_time = current_datetime.strftime("%H:%M:%S")
    location = "Gurugram"

    text_info.append((f"{formatted_date}", (410, 550)))
    text_info.append((f"{formatted_time}", (410, 620)))
    text_info.append(("Rs.150", (410, 690)))
    text_info.append((f"{location}", (410, 760)))

    # Create the canvas
    ticket_width, ticket_height = ticket_template.size
    total_width = ticket_width
    total_height = number * (ticket_height + margin)
    canvas = Image.new('RGB', (total_width, total_height), color='white')

    y_offset = 0
    
    for i in range(start_number,number + start_number):
        # Make a copy of the ticket template for each ticket
        ticket_template_copy = ticket_template.copy()

        qr = qrcode.QRCode(
            version=1,
            box_size=15,
            border=4,
        )
        qr_data = f'{data_prefix}_{i}'
        qr.add_data(qr_data)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")

        # Paste the QR code onto the ticket template copy    
        qr_width, qr_height = qr_img.size
        qr_x_offset = ticket_width // 2
        qr_y_offset = y_offset
        ticket_template_copy.paste(qr_img, (qr_x_offset, qr_y_offset))

        # Draw text on the ticket template copy
        draw = ImageDraw.Draw(ticket_template_copy)
        font = ImageFont.truetype(font_path, size=font_size)
        for text, position in text_info:
            draw.text(position, text, fill=color, font=font)

        # Add ticket number to text information
        ticket_number = f"Ticket Number: {i}"
        draw.text((210, 490), ticket_number, fill=color, font=font)

        # Paste the ticket template copy onto the canvas
        canvas.paste(ticket_template_copy, (0, (ticket_height + margin) * (i - start_number)))

    # Return the canvas
    return canvas





import os.path

COUNTERS_FILE_PATH = 'JKMapp/static/counter_info.xlsx'

def read_counters():
    total_count = 0
    try:
        if os.path.exists(COUNTERS_FILE_PATH):
            workbook = load_workbook(COUNTERS_FILE_PATH)
            sheet = workbook.active
            total_count = sheet.cell(row=1, column=5).value or 0
        else:
            workbook = Workbook()
            sheet = workbook.active
            workbook.save(COUNTERS_FILE_PATH)
    except Exception as e:
        print(f"Error reading counters: {e}")
    return total_count

def update_counters(total_count, ticket_count):
    try:
        workbook = load_workbook(COUNTERS_FILE_PATH) if os.path.exists(COUNTERS_FILE_PATH) else Workbook()
        sheet = workbook.active

        row = sheet.max_row + 1
        current_datetime = datetime.datetime.utcnow() + datetime.timedelta(hours=5, minutes=30)
        formatted_date = current_datetime.strftime("%d-%m-%Y")
        formatted_time = current_datetime.strftime("%H:%M:%S")
        
        total_tickets = ticket_count

        current_iteration = sheet.cell(row=row-1, column=1).value or 0
        current_iteration += 1

        sheet.cell(row=row, column=1).value = current_iteration
        sheet.cell(row=row, column=2).value = ticket_count
        sheet.cell(row=row, column=3).value = formatted_date
        sheet.cell(row=row, column=4).value = formatted_time

        total_tickets_cell = sheet.cell(row=1, column=5)
        total_tickets_cell.value = total_tickets_cell.value or 0
        total_tickets_cell.value += total_tickets
        
        workbook.save(COUNTERS_FILE_PATH)
    except Exception as e:
        print(f"Error updating counters: {e}")



def generate_pdf_from_excel(input_excel, output_pdf):
    # Load the Excel file
    wb = load_workbook(input_excel)
    ws = wb.active

    # Create a PDF
    c = canvas.Canvas(output_pdf, pagesize=letter)
    y = 700  # Initial y-coordinate for writing text

    for row in ws.iter_rows():
        x = 50  # Initial x-coordinate for writing text
        for cell in row:
            c.drawString(x, y, str(cell.value))
            x += 100  # Move to the next column
        y -= 12  # Move to the next row
    c.save()


















def BarcodeReader(image_path):
    # Read the image
    img = cv2.imread(image_path)
      
    # Decode the barcode image
    detectedBarcodes = decode(img)
      
    # If not detected then print the message
    if not detectedBarcodes:
        print("QR Code Not Detected or your QR code is blank/corrupted!")
    else:
        # Traverse through all the detected barcodes in the image
        for barcode in detectedBarcodes:
            # Print the barcode data
            print("QR Code Data:", barcode.data)
        return barcode.data
    

import csv

def check_and_save_string(input_string, file_path):
    # Read existing strings from the CSV file
    existing_strings = set()
    try:
        with open(file_path, 'r') as file:
            reader = csv.reader(file)
            for row in reader:
                existing_strings.add(row[0])  # Assuming each row contains a single string
    except FileNotFoundError:
        pass  # File not found, assume no existing strings

    # Check if input_string already exists
    if input_string.decode('utf-8') in existing_strings:
        print("String is repeated")
    else:
        # Save the string to the CSV file
        with open(file_path, 'a', newline='') as file:
            writer = csv.writer(file)
            writer.writerow([input_string.decode('utf-8')])
        print("String saved successfully")

# Example usage:
#input_string = b'JKM2024_1'
#check_and_save_string(input_string, file_path)
